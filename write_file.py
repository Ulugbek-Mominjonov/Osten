import string

from PIL import Image as Images
from PIL import ImageFont
from PIL import ImageDraw

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

def write_img(data):
    with Images.open("rasm/maket.jpg").convert("RGBA") as base:
        txt = Images.new("RGBA", base.size, (0, 0, 0, 0))

        d = ImageDraw.Draw(txt)

        fnt = ImageFont.truetype("shrift/Roboto-Regular.ttf", 8)

        # draw text, half opacity
        text = "O'lchovchining (F.I.Sh.)"
        d.text((127, 96), f"{data['Loyiha_nomi']}", font=fnt, fill=(0, 0, 0))
        d.text((127, 106), f"{data['Sana']}", font=fnt, fill=(0, 0, 0))
        d.text((127, 116), f"{data[f'{text}']}", font=fnt, fill=(0, 0, 0))
        if data['Shahta_turi'] == "Beton":
            d.text((125, 126), "+", font=fnt, fill=(0, 0, 0))
        elif data['Shahta_turi'] == "Bloklar":
            d.text((184, 126), "+", font=fnt, fill=(0, 0, 0))
        else:
            d.text((238, 126), "+", font=fnt, fill=(0, 0, 0))

        if data['Mrl_Or_Rl'] == "MR":
            d.text((173, 146), "+", font=fnt, fill=(0, 0, 0))
        else:
            d.text((218, 146), "+", font=fnt, fill=(0, 0, 0))

        d.text((313, 190), f"{data['olchamlar']['A']}", font=fnt, fill=(0, 0, 0))
        d.text((313, 200), f"{data['olchamlar']['B']}", font=fnt, fill=(0, 0, 0))
        d.text((313, 210), f"{data['olchamlar']['C']}", font=fnt, fill=(0, 0, 0))
        d.text((313, 220), f"{data['olchamlar']['D']}", font=fnt, fill=(0, 0, 0))
        d.text((313, 230), f"{data['olchamlar']['E']}", font=fnt, fill=(0, 0, 0))

        i = 292
        for item in data['qavatlar']:
            d.text((313, i), f"{data['qavatlar'][f'{item}']}", font=fnt, fill=(0, 0, 0))
            i += 10

        d.text((313, 455), f"{len(data['qavatlar'])}", font=fnt, fill=(0, 0, 0))
        d.text((313, 465), f"{data['eshiklar_soni']}", font=fnt, fill=(0, 0, 0))
        d.text((313, 475), f"{data['P']}", font=fnt, fill=(0, 0, 0))
        d.text((313, 485), f"{data['Hp']}", font=fnt, fill=(0, 0, 0))
        d.text((261, 516), f"{data['Eslatma']}", font=fnt, fill=(0, 0, 0))

        out = Images.alpha_composite(base, txt)

        out.save("salom.png")

def write_excel_osten(keys, data):
    wb = Workbook()
    dest_filename = f"{data['Loyiha_nomi']}.xlsx"
    ws1 = wb.active
    ws1.title = "Olchamlar"
    ws1.sheet_properties.tabColor = "1072BA"

    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 25
    ws1.column_dimensions["C"].width = 5
    ws1.column_dimensions["D"].width = 20
    ws1.column_dimensions["E"].width = 25
    ws1.column_dimensions["F"].width = 30
    ws1.column_dimensions["G"].width = 20

    # logo
    ws1.cell(row=1, column=2).border = thin_border
    img = Image('rasm/logo.png')
    img.width = 160
    img.height = 70
    ws1.cell(row=1, column=1).border = thin_border
    ws1.merge_cells('A1:C5')
    ws1.add_image(img, 'B2')

    cell = ws1.cell(row=1, column=4)
    cell.value = "O'lcham loyihalari jadvali"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=1, column=4).border = thin_border
    ws1["D1"].font = Font(bold=True)
    ws1.merge_cells('D1:F5')

    for i in range(6, 11):
        ws1.cell(row=i, column=1).border = thin_border
        ws1.cell(row=i, column=3).border = thin_border
        ws1.cell(row=i, column=4).border = thin_border
    ws1.merge_cells('A6:F6')
    ws1.cell(row=7, column=1, value=keys[0])
    ws1.merge_cells('A7:B7')
    ws1.merge_cells('C7:F7')
    ws1.cell(row=7, column=3, value=data['Loyiha_nomi'])
    ws1["C7"].font = Font(bold=True)

    ws1.cell(row=8, column=1, value=keys[1])
    ws1.merge_cells('A8:B8')
    ws1.merge_cells('C8:F8')
    ws1.cell(row=8, column=3, value=data[f'{keys[1]}'])

    ws1.cell(row=9, column=1, value=keys[2])
    ws1.merge_cells('A9:B9')
    ws1.merge_cells('C9:F9')
    ws1.cell(row=9, column=3, value=data[f'{keys[2]}'])

    ws1.cell(row=10, column=1, value=keys[3])
    ws1.merge_cells('A10:B10')
    shahta = ["Beton", "Bloklar", "Po'lat Konstruksiya"]
    if data['Shahta_turi'] == "Beton":
        text = f"+ {data['Shahta_turi']}"
        ws1.cell(row=10, column=4, value=text).alignment = Alignment(horizontal='center', vertical='center')
        ws1["D10"].font = Font(bold=True)
        ws1.cell(row=10, column=5, value=shahta[1]).alignment = Alignment(horizontal='center', vertical='center')
        ws1.cell(row=10, column=6, value=shahta[2]).alignment = Alignment(horizontal='center', vertical='center')
    elif data['Shahta_turi'] == "Bloklar":
        text = f"+ {data['Shahta_turi']}"
        ws1.cell(row=10, column=5, value=text).alignment = Alignment(horizontal='center', vertical='center')
        ws1["E10"].font = Font(bold=True)
        ws1.cell(row=10, column=4, value=shahta[0]).alignment = Alignment(horizontal='center', vertical='center')
        ws1.cell(row=10, column=6, value=shahta[2]).alignment = Alignment(horizontal='center', vertical='center')
    else:
        text = f"+ {data['Shahta_turi']}"
        ws1.cell(row=10, column=6, value=text).alignment = Alignment(horizontal='center', vertical='center')
        ws1["F10"].font = Font(bold=True)
        ws1.cell(row=10, column=4, value=shahta[0]).alignment = Alignment(horizontal='center', vertical='center')
        ws1.cell(row=10, column=5, value=shahta[1]).alignment = Alignment(horizontal='center', vertical='center')

    for i in range(5, 7):
        ws1.cell(row=10, column=i).border = thin_border

    MRL_MR = ["MR", "MRL"]
    if data['Mrl_Or_Rl'] == "MR":
        text = f"+ {data['Mrl_Or_Rl']}"
        ws1.cell(row=11, column=1, value=text).alignment = Alignment(horizontal='center', vertical='center')
        ws1["A11"].font = Font(bold=True)
        ws1.cell(row=11, column=5, value=MRL_MR[1]).alignment = Alignment(horizontal='center', vertical='center')
    elif data['Mrl_Or_Rl'] == "MRL":
        text = f"+ {data['Mrl_Or_Rl']}"
        ws1.cell(row=11, column=1, value=MRL_MR[0]).alignment = Alignment(horizontal='center', vertical='center')
        ws1.cell(row=11, column=5, value=text).alignment = Alignment(horizontal='center', vertical='center')
        ws1["E11"].font = Font(bold=True)

    ws1.cell(row=11, column=1).border = thin_border
    ws1.cell(row=11, column=5).border = thin_border

    ws1.merge_cells('A11:D11')
    ws1.merge_cells('E11:F11')

    img = Image("rasm/rasm_1.jpg")
    img.width = 290
    img.height = 290
    ws1.add_image(img, 'B13')
    img = Image("rasm/rasm_2.jpg")
    img.width = 270
    img.height = 430
    ws1.cell(row=12, column=1).border = thin_border
    ws1.merge_cells('A12:D56')
    ws1.add_image(img, 'B30')


    text = "O'lcham(mm)"
    ws1.cell(row=12, column=6, value=text).alignment = Alignment(horizontal='center', vertical='center')
    ws1["F12"].font = Font(bold=True)
    ws1.cell(row=12, column=5, value=f"Shartli belgilar").alignment = Alignment(horizontal='center', vertical='center')
    ws1["E12"].font = Font(bold=True)
    ws1.cell(row=12, column=5).border = thin_border
    ws1.cell(row=12, column=6).border = thin_border

    olcham = []
    for item in data['olchamlar']:
        olcham.append(item)
    ws1.cell(row=13, column=5, value=olcham[0]).alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=13, column=6, value=data["olchamlar"]["A"]).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')

    ws1.cell(row=14, column=5, value=olcham[1]).alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=14, column=6, value=data["olchamlar"]["B"]).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')

    ws1.cell(row=15, column=5, value=olcham[2]).alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=15, column=6, value=data["olchamlar"]["C"]).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')

    ws1.cell(row=16, column=5, value=olcham[3]).alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=16, column=6, value=data["olchamlar"]["D"]).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')

    ws1.cell(row=17, column=5, value=olcham[4]).alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=17, column=6, value=data["olchamlar"]["E"]).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')
    for i in range(13, 18):
        ws1.cell(row=i, column=5).border = thin_border
        ws1.cell(row=i, column=6).border = thin_border

    ws1.cell(row=18, column=5).border = thin_border
    ws1.merge_cells('E18:F19')

    ws1.cell(row=20, column=5, value=f'Qavat №').alignment = Alignment(horizontal='center', vertical='center')
    ws1["E20"].font = Font(bold=True)
    ws1.cell(row=20, column=6, value=f'Balandlik(H) (mm)').alignment = Alignment(horizontal='center', vertical='center')
    ws1["F20"].font = Font(bold=True)
    ws1.cell(row=20, column=5).border = thin_border
    ws1.cell(row=20, column=6).border = thin_border
    i = 21
    a = 0
    for item, value in data['qavatlar'].items():
        ws1.cell(row=i, column=5, value=item).alignment = Alignment(horizontal='center', vertical='center')
        ws1.cell(row=i, column=6, value=value).alignment = Alignment(horizontal='center', vertical='center')
        ws1.cell(row=i, column=5).border = thin_border
        ws1.cell(row=i, column=6).border = thin_border
        i += 1
        a = item

    q= int(a)+1
    for j in range(i, i+20-int(a)):
        ws1.cell(row=i, column=5, value=q).alignment = Alignment(horizontal='center', vertical='center')
        ws1.cell(row=i, column=5).border = thin_border
        ws1.cell(row=i, column=6).border = thin_border
        i += 1
        q+=1

    ws1.cell(row=41, column=5).border = thin_border
    ws1.merge_cells('E41:F43')

    ws1.cell(row=44, column=5, value=f'Umumiy qavatlar soni').alignment = Alignment(horizontal='center', vertical='center')
    ws1["E44"].font = Font(bold=True)
    ws1.cell(row=44, column=6, value=len(data['qavatlar'])).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')

    ws1.cell(row=45, column=5, value=f'Eshiklar soni').alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=45, column=6, value=data['eshiklar_soni']).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')

    ws1.cell(row=46, column=5, value=f'Priyamka balandligi(P)').alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=46, column=6, value=data['P']).alignment = Alignment(horizontal='center', vertical='center')

    ws1.cell(row=47, column=5, value=f'Oxirgi etaj balandligi(H)').alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=47, column=6, value=data['Hp']).alignment = Alignment(horizontal='center', vertical='center')

    ws1.cell(row=48, column=5, value=f'Eslatma').alignment = Alignment(horizontal='center', vertical='center')
    ws1["E48"].font = Font(bold=True)
    ws1.cell(row=48, column=6, value=data['Eslatma']).alignment = Alignment(horizontal='center', vertical='center', wrapText=True )

    for i in range(44, 49):
        ws1.cell(row=i, column=5).border = thin_border
        ws1.cell(row=i, column=6).border = thin_border
    ws1.merge_cells('E48:E56')
    ws1.merge_cells('F48:F56')
    wb.save(filename=dest_filename)
    print("ishladi")

def write_excel_birnima(data):
    wb = Workbook()
    dest_filename = f"{data['Лойиҳа номи']}.xlsx"
    ws2 = wb.active
    ws2.title = "Olchamlar"
    ws2.sheet_properties.tabColor = "1072BA"

    # ws2.column_dimensions["A"].width = 15
    # ws2.column_dimensions["B"].width = 15
    # ws2.column_dimensions["C"].width = 15
    # ws2.column_dimensions["D"].width = 15

    # logo
    img = Image('rasm/rasm_2_bot.png')
    ws2.merge_cells('H2:N22')
    ws2.add_image(img, 'H2')

    i = 2
    for item in data:
        if item != 'state':
            ws2.cell(row=i, column=1, value=f'{item}:')
            ws2.cell(row=i, column=4, value=f"{data[f'{item}']}").alignment = Alignment(horizontal='center', vertical='center')
            ws2.cell(row=i, column=1).border = thin_border
            ws2.cell(row=i, column=4).border = thin_border

            ws2.merge_cells(f'A{i}:C{i}')
            ws2.merge_cells(f'D{i}:F{i}')
            i += 1
        else:
            break
    qavat_keys = ["A", "B", "C", "D", "E", "F"]
    d = 1
    e = 2
    for j in range(1, 10):
        if len(data[f'qavat_{j}']) != 0:
            ws2.cell(row=24, column=d, value=f'№')
            ws2.cell(row=24, column=e, value=f"Этаж - {j}").alignment = Alignment(horizontal='center',vertical='center')
            ws2.cell(row=24, column=d).border = thin_border
            ws2.cell(row=24, column=e).border = thin_border

            for i in range(1, 7):
                ws2.cell(row=24 + i, column=d, value=f'{qavat_keys[i - 1]}')
                son = data[f'qavat_{j}'][qavat_keys[i - 1]]
                ws2.cell(row=24 + i, column=e, value=son).alignment = Alignment(horizontal='center', vertical='center')
                ws2.cell(row=24 + i, column=d).border = thin_border
                ws2.cell(row=24 + i, column=e).border = thin_border

        d += 2
        e += 2

    d = 1
    e = 2
    for j in range(10, 19):
        if len(data[f'qavat_{j}']) != 0:
            ws2.cell(row=32, column=d, value=f'№')
            ws2.cell(row=32, column=e, value=f"Этаж - {j}").alignment = Alignment(horizontal='center',vertical='center')
            ws2.cell(row=32, column=d).border = thin_border
            ws2.cell(row=32, column=e).border = thin_border

            for i in range(1, 7):
                ws2.cell(row=32 + i, column=d, value=f'{qavat_keys[i - 1]}')
                son = data[f'qavat_{j}'][qavat_keys[i - 1]]
                ws2.cell(row=32 + i, column=e, value=son).alignment = Alignment(horizontal='center', vertical='center')
                ws2.cell(row=32 + i, column=d).border = thin_border
                ws2.cell(row=32 + i, column=e).border = thin_border
        d += 2
        e += 2
    wb.save(filename=dest_filename)

def write_excel_osten3(data):
    wb = Workbook()
    dest_filename = f"{data['Loyiha nomi']}.xlsx"
    ws2 = wb.active
    ws2.title = "Olchamlar"
    ws2.sheet_properties.tabColor = "1072BA"

    keys = []
    for item in data.keys():
        keys.append(item)
    ws2.column_dimensions["A"].width = 3
    for i in list(string.ascii_uppercase)[1:12]:
        ws2.column_dimensions[f"{i}"].width = 9

    for i in list(string.ascii_uppercase)[16:26]:
        ws2.column_dimensions[f"{i}"].width = 7

    ws2.column_dimensions["M"].width = 11
    ws2.column_dimensions["N"].width = 11
    ws2.column_dimensions["O"].width = 7
    ws2.column_dimensions["P"].width = 3

    ws2.cell(row=2, column=2, value=f"Lift shaxtasi haqida ma'lumot").alignment = Alignment(horizontal='center', vertical='center')
    ws2["B2"].font = Font(bold=True)
    ws2.cell(row=2, column=2).border = thin_border
    ws2.merge_cells('B2:D4')

    ws2.cell(row=2, column=5, value=f"Loyiha nomi").alignment = Alignment(horizontal='center', vertical='center')
    ws2["E2"].font = Font(bold=True)
    ws2.cell(row=2, column=5).border = thin_border
    ws2.merge_cells('E2:F4')

    ws2.cell(row=2, column=7, value=f"{data[keys[0]]}").alignment = Alignment(horizontal='center', vertical='center')
    ws2["G2"].font = Font(bold=True)
    ws2.cell(row=2, column=7).border = thin_border
    ws2.merge_cells('G2:N4')


    kalit = ['Qavat №', 'Qavat nomi', 'Chap A', 'Chap B', 'Chap C', "O'ng D", "O'ng E", "O'ng F", 'G', 'I', 'J', 'Qavat baladligi(H)', 'Eshik baladligi(K)']
    for i in range(25, 5, -1):
        for j in range(2, 15):
            ws2.cell(row=i, column=j).border = thin_border
            ws2.cell(row=i, column=j).alignment = Alignment(horizontal='center', vertical='center')

    j = 2
    for i in kalit:
        ws2.cell(row=5, column=j, value=f'{i}').alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        ws2.cell(row=5, column=j).font = Font(bold=True)
        ws2.cell(row=5, column=j).border = thin_border
        j+=1

    r = 25
    for i in range(1, 21):
        c = 4
        if len(data[f"qavat_{i}"]) != 0:
            for j in data[f"qavat_1"]:
                q = data[f"qavat_{i}"][j]
                ws2.cell(row=r, column=c, value=f"{q}")
                c += 1
        ws2.cell(row=r, column=2, value=f'{i}')
        r -= 1

    ws2.cell(row=26, column=2, value=f'Avtomatik tanlash').alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws2.cell(row=26, column=3).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws2.cell(row=27, column=2, value=f"Qo'lda tanlash").alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws2.cell(row=26, column=2).border = thin_border
    ws2.cell(row=27, column=2).border = thin_border
    ws2.merge_cells('B26:C26')
    ws2.merge_cells('B27:C27')

    l = 4
    arr = []
    for i in kalit[2:]:
        if i != 'Qavat baladligi(H)':
            for j in range(1, 21):
                if len(data[f"qavat_{j}"]) != 0:
                    arr.append(data[f"qavat_{j}"][i])
            if i == 'I':
                maxx = max(arr)
                ws2.cell(row=26, column=l, value=maxx).alignment = Alignment(horizontal='center', vertical='center',
                                                                      wrapText=True)
            else:
                minn = min(arr)
                ws2.cell(row=26, column=l, value=minn).alignment = Alignment(horizontal='center', vertical='center',
                                                                             wrapText=True)
        else:
            ws2.cell(row=26, column=l, value=f"").alignment = Alignment(horizontal='center', vertical='center',
                                                                         wrapText=True)

        ws2.cell(row=26, column=l).border = thin_border
        ws2.cell(row=27, column=l).border = thin_border
        l += 1
        arr = []

    ws2.cell(row=28, column=2, value=f"Ishlatiladigan rele o'lchov birligi mm bo'lishi kerak.").alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws2.cell(row=28, column=2).font = Font(color="FFFF0000")
    ws2.cell(row=28, column=2).border = thin_border
    ws2.cell(row=29, column=2, value=f"Agarda shaxtada profil mavjud bo'lsa, uning o'lchamlari qo'shimcha varaqda ko'rsatilishi kerak.").alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws2.cell(row=29, column=2).font = Font(color="FFFF0000")
    ws2.cell(row=29, column=2).border = thin_border
    ws2.merge_cells('B28:N28')
    ws2.merge_cells('B29:N29')

    ws2.cell(row=31, column=2, value=f"Sim o'rtasi").alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws2.cell(row=31, column=2).font = Font(bold=True)
    ws2.cell(row=31, column=2).border = thin_border
    ws2.cell(row=31, column=5, value=data["Sim o'rtasi"]).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws2.cell(row=31, column=5).border = thin_border
    ws2.merge_cells('B31:C31')
    ws2.merge_cells('E31:F31')

    ws2.cell(row=31, column=8, value=f"Devor qalinligi").alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws2.cell(row=31, column=8).font = Font(bold=True)
    ws2.cell(row=31, column=8).border = thin_border
    ws2.cell(row=31, column=12, value=data["Devor qalinligi"]).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws2.cell(row=31, column=12).border = thin_border
    ws2.merge_cells('H31:J31')
    ws2.merge_cells('L31:M31')

    ##

    ws2.cell(row=34, column=2, value=f"Shaxta kengligi").alignment = Alignment(horizontal='center', vertical='center',
                                                                           wrapText=True)
    ws2.cell(row=34, column=2).font = Font(bold=True)
    ws2.cell(row=34, column=2).border = thin_border
    foo = int(ws2["D26"].value) + int(ws2["G26"].value) + int(ws2["E31"].value)
    ws2.cell(row=34, column=5, value=foo).alignment = Alignment(horizontal='center', vertical='center',
                                                                                wrapText=True)
    ws2.cell(row=34, column=5).border = thin_border
    ws2.merge_cells('B34:C34')
    ws2.merge_cells('E34:F34')

    ws2.cell(row=34, column=8, value=f"Shaxtaning chuqurligi").alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws2.cell(row=34, column=8).font = Font(bold=True)
    ws2.cell(row=34, column=8).border = thin_border
    foo = min(int(ws2["E26"].value) + int(ws2["F26"].value), int(ws2["H26"].value) + int(ws2["I26"].value))
    ws2.cell(row=34, column=12, value=foo).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws2.cell(row=34, column=12).border = thin_border
    ws2.merge_cells('H34:J34')
    ws2.merge_cells('L34:M34')

    ##
    ws2.cell(row=37, column=2, value=f"Priyamka chuqurligi").alignment = Alignment(horizontal='center', vertical='center',
                                                                               wrapText=True)
    ws2.cell(row=37, column=2).font = Font(bold=True)
    ws2.cell(row=37, column=2).border = thin_border
    ws2.cell(row=37, column=5, value=data["Priyamka chuqurligi"]).alignment = Alignment(horizontal='center', vertical='center',
                                                                wrapText=True)
    ws2.cell(row=37, column=5).border = thin_border
    ws2.merge_cells('B37:C37')
    ws2.merge_cells('E37:F37')

    ws2.cell(row=37, column=8, value=f"Oxirigi qavatning balandligi").alignment = Alignment(horizontal='center',
                                                                                     vertical='center', wrapText=True)
    ws2.cell(row=37, column=8).font = Font(bold=True)
    ws2.cell(row=37, column=8).border = thin_border
    ws2.cell(row=37, column=12, value=data["Oxirgi qavat balandligi"]).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws2.cell(row=37, column=12).border = thin_border
    ws2.merge_cells('H37:J37')
    ws2.merge_cells('L37:M37')

    ##
    ws2.cell(row=40, column=2, value=f"Priyamka devor turi").alignment = Alignment(horizontal='center',
                                                                                   vertical='center',
                                                                                   wrapText=True)
    ws2.cell(row=40, column=2).font = Font(bold=True)
    ws2.cell(row=40, column=2).border = thin_border
    ws2.cell(row=40, column=5, value=data["Priyamka devor turi"]).alignment = Alignment(horizontal='center',
                                                                                        vertical='center',
                                                                                        wrapText=True)
    ws2.cell(row=40, column=5).border = thin_border
    ws2.merge_cells('B40:C40')
    ws2.merge_cells('E40:F40')

    ws2.cell(row=40, column=8, value=f"Priyamka tagida bo'sh devor bormi?").alignment = Alignment(horizontal='center',
                                                                                            vertical='center',
                                                                                            wrapText=True)
    ws2.cell(row=40, column=8).font = Font(bold=True)
    ws2.cell(row=40, column=8).border = thin_border
    ws2.cell(row=40, column=14, value=data["Priyamka tagida bo'sh devor bormi?"]).alignment = Alignment(horizontal='center',
                                                                                             vertical='center',
                                                                                             wrapText=True)
    ws2.cell(row=40, column=14).border = thin_border
    ws2.merge_cells('H40:M40')

    # CHIZMA
    ws2.cell(row=5, column=17, value=data["orqa devor"]).alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=5, column=17).font = Font(bold=True)
    ws2.cell(row=5, column=17).border = Border(left=Side(style='double'),
                     right=Side(style='double'),
                     top=Side(style='double'),
                     bottom=Side(style='double'))
    ws2.merge_cells('Q5:Z5')

    ws2.cell(row=6, column=17, value=data["chap_yon devor"]).alignment = Alignment(
        horizontal='center',
        vertical='center',
        textRotation=90,
        wrapText=True)
    ws2.cell(row=6, column=17).font = Font(bold=True)
    ws2.cell(row=6, column=17).border = Border(left=Side(style='double'),
                     right=Side(style='double'),
                     top=Side(style='double'),
                     bottom=Side(style='double'))
    ws2.merge_cells('Q6:Q22')

    ws2.cell(row=6, column=26, value=data["ong_yon devor"]).alignment = Alignment(
        horizontal='center',
        vertical='center',
        textRotation=90,
        wrapText=True)
    ws2.cell(row=6, column=26).font = Font(bold=True)
    ws2.cell(row=6, column=26).border = Border(left=Side(style='double'),
                     right=Side(style='double'),
                     top=Side(style='double'),
                     bottom=Side(style='double'))
    ws2.merge_cells('Z6:Z22')

    ws2.cell(row=21, column=18, value=data["old_chap devor"]).alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=21, column=18).font = Font(bold=True)
    ws2.cell(row=21, column=18).border = Border(left=Side(style='double'),
                     right=Side(style='double'),
                     top=Side(style='double'),
                     bottom=Side(style='double'))
    ws2.merge_cells('R21:T22')

    ws2.cell(row=21, column=23, value=data["old_ong devor"]).alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=21, column=23).font = Font(bold=True)
    ws2.cell(row=21, column=23).border = Border(left=Side(style='double'),
                     right=Side(style='double'),
                     top=Side(style='double'),
                     bottom=Side(style='double'))
    ws2.merge_cells('W21:Y22')

    ws2.cell(row=6, column=19, value=f"Chap orqa C").alignment = Alignment(
        horizontal='right',
        vertical='center',
        textRotation=90,
        wrapText=True)
    ws2.cell(row=6, column=19).border = Border(right=Side(style='thin'))
    ws2.merge_cells('S6:S12')

    ws2.cell(row=13, column=18, value=f"Chap A").alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=13, column=18).border = thin_border
    ws2.merge_cells('R13:S13')

    ws2.cell(row=14, column=19, value=f"Chap oldi B").alignment = Alignment(
        horizontal='right',
        vertical='center',
        textRotation=90,
        wrapText=True)
    ws2.cell(row=14, column=19).border = Border(right=Side(style='thin'))
    ws2.merge_cells('S14:S20')
    #
    ws2.cell(row=13, column=21, value=f"Sim o'rtasi").alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=13, column=21).border = thin_border
    ws2.merge_cells('U13:V13')
    #
    ws2.cell(row=6, column=24, value=f"O'ng orqa F").alignment = Alignment(
        horizontal='right',
        vertical='center',
        textRotation=90,
        wrapText=True)
    ws2.cell(row=6, column=24).border = Border(right=Side(style='thin'))
    ws2.merge_cells('X6:X12')

    ws2.cell(row=13, column=24, value=f"O'ng D").alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=13, column=24).border = thin_border
    ws2.merge_cells('X13:Y13')

    ws2.cell(row=14, column=24, value=f"O'ng oldi E").alignment = Alignment(
        horizontal='right',
        vertical='center',
        textRotation=90,
        wrapText=True)
    ws2.cell(row=14, column=24).border = Border(right=Side(style='thin'))
    ws2.merge_cells('X14:X20')

    ws2.cell(row=23, column=18, value=f"Eshik chap G").alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=23, column=18).border = Border(bottom=Side(style='thin'))
    ws2.merge_cells('R23:S23')

    ws2.cell(row=23, column=21, value=f"Eshik orasi I").alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=23, column=21).border = Border(bottom=Side(style='thin'))
    ws2.merge_cells('U23:V23')

    ws2.cell(row=23, column=24, value=f"Eshik o'ng J").alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=23, column=24).border = Border(bottom=Side(style='thin'))
    ws2.merge_cells('X23:Y23')

    ws2.cell(row=25, column=17, value=f"Yuqoridagi maydonda siz devorining turini ko'rsatishingiz kerak.").alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=25, column=17).font = Font(color="FFFF0000")
    ws2.cell(row=25, column=17).border = thin_border
    ws2.merge_cells('Q25:Z25')

    ws2.cell(row=30, column=16, value=f"Dvigatel xonasi").alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=30, column=16).font = Font(bold=True)
    ws2.cell(row=30, column=16).border = thin_border
    ws2.merge_cells('P30:Z30')

    ws2.cell(row=31, column=16, value=f"Stanina ostidagi betonning balandligi:").alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=31, column=16).border = thin_border
    ws2.merge_cells('P31:S31')

    ws2.cell(row=31, column=20, value=data["Stanina ostidagi betonning balandligi"]).alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=31, column=20).border = thin_border

    ws2.cell(row=31, column=21, value=f"Mashina xonasi balandligi").alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=31, column=21).border = thin_border
    ws2.merge_cells('U31:Y31')

    ws2.cell(row=31, column=26, value=data["Mashina xonasi balandligi"]).alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=31, column=26).border = thin_border

    ws2.cell(row=34, column=16, value=f"Eslatma: {data['Eslatma']}").alignment = Alignment(
        horizontal='left',
        vertical='center',
        wrapText=True)
    ws2.cell(row=34, column=16).font = Font(bold=True)
    ws2.cell(row=34, column=16).border = thin_border
    ws2.merge_cells('P34:Z37')

    ws2.cell(row=39, column=16, value=f'Muhandis').alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=39, column=16).font = Font(bold=True)
    ws2.cell(row=39, column=16).border = thin_border
    ws2.merge_cells('P39:S39')

    ws2.cell(row=40, column=16, value=data['Muhandis']).alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=40, column=16).border = thin_border
    ws2.merge_cells('P40:S40')
    #
    ws2.cell(row=39, column=21, value=f'Sana').alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=39, column=21).font = Font(bold=True)
    ws2.cell(row=39, column=21).border = thin_border
    ws2.merge_cells('U39:V39')

    ws2.cell(row=40, column=21, value=data['Sana']).alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=40, column=21).border = thin_border
    ws2.merge_cells('U40:V40')

    #
    ws2.cell(row=39, column=24, value=f'Imzo').alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=39, column=24).font = Font(bold=True)
    ws2.cell(row=39, column=24).border = thin_border
    ws2.merge_cells('X39:Y39')

    ws2.cell(row=40, column=24).alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True)
    ws2.cell(row=40, column=24).border = thin_border
    ws2.merge_cells('X40:Y40')

    wb.save(filename=dest_filename)
